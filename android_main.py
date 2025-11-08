#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
虫害情况自动提取工具 - Android版本 v2.0
使用 Kivy 框架开发的移动端界面
"""

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.progressbar import ProgressBar
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.metrics import dp
from pathlib import Path
import threading
import re
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

# Android 存储路径配置
try:
    from android.permissions import request_permissions, Permission, check_permission
    from android.storage import app_storage_path, primary_external_storage_path
    from jnius import autoclass, cast
    ANDROID = True
    
    # 获取应用存储路径
    try:
        PythonActivity = autoclass('org.kivy.android.PythonActivity')
        Environment = autoclass('android.os.Environment')
        context = PythonActivity.mActivity
        
        # 优先使用 Documents 目录
        documents_dir = Environment.DIRECTORY_DOCUMENTS
        external_files = context.getExternalFilesDir(documents_dir)
        
        if external_files:
            STORAGE_PATH = str(external_files.getAbsolutePath())
            print(f"使用应用专属外部存储: {STORAGE_PATH}")
        else:
            # Fallback: 应用内部存储
            STORAGE_PATH = str(app_storage_path())
            print(f"使用应用内部存储: {STORAGE_PATH}")
            
    except Exception as e:
        print(f"警告: 存储路径获取失败: {e}")
        try:
            STORAGE_PATH = str(app_storage_path())
        except:
            STORAGE_PATH = "/data/data/com.pestcontrol.pestreportextractor/files"
            
except ImportError:
    ANDROID = False
    STORAGE_PATH = str(Path.home())


class PestReportExtractor:
    """虫害报告提取器核心类"""
    
    def __init__(self):
        self.pdf_path = None
        self.pest_data = []
        self.output_path = None
        
    def extract_pest_data_from_pdf(self, pdf_path):
        """从PDF中提取虫害数据"""
        self.pest_data = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    full_text += page.extract_text() + "\n"
                
                pest_section = self._extract_pest_section(full_text)
                
                if not pest_section:
                    return False, "未找到虫害情况数据"
                
                self._parse_pest_records(pest_section)
                
                return True, f"成功提取 {len(self.pest_data)} 条记录"
                
        except Exception as e:
            return False, f"PDF读取失败: {str(e)}"
    
    def _extract_pest_section(self, text):
        """提取虫害情况部分的文本"""
        start_marker = "虫害情况"
        end_marker = "服务总结"
        
        start_idx = text.find(start_marker)
        end_idx = text.find(end_marker)
        
        if start_idx != -1 and end_idx != -1:
            return text[start_idx:end_idx]
        return None
    
    def _parse_pest_records(self, text):
        """解析虫害记录"""
        text = text.replace('\x01', ' ')
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            pest_match = re.match(r'^([\u4e00-\u9fa5]+)\s+发现虫害活动\s*[-–—]\s*(\d+)', line)
            if pest_match:
                pest_type = pest_match.group(1)
                count = int(pest_match.group(2))
                
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    building_match = re.search(
                        r'建筑物:\s*([^,]+),\s*楼层:\s*([^,]+),\s*部门:\s*([^,]+),\s*检查/发现监测点位:\s*(.+)',
                        next_line
                    )
                    
                    if building_match:
                        self.pest_data.append({
                            '建筑物': building_match.group(1).strip(),
                            '楼层': building_match.group(2).strip(),
                            '部门': building_match.group(3).strip(),
                            '检查/发现监测点位': building_match.group(4).strip(),
                            '虫害类型': pest_type,
                            '发现虫害活动': count
                        })
    
    def create_excel(self, output_dir=None):
        """创建Excel文件"""
        if not self.pest_data:
            return False, "没有数据可以导出"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "虫害情况"
        
        headers = ["建筑物", "楼层", "部门", "检查/发现监测点位", "虫害类型", "发现虫害活动"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for record in self.pest_data:
            ws.append([
                record['建筑物'],
                record['楼层'],
                record['部门'],
                record['检查/发现监测点位'],
                record['虫害类型'],
                record['发现虫害活动']
            ])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        if output_dir is None:
            output_dir = Path(STORAGE_PATH) / "Documents" / "虫害报告"
        else:
            output_dir = Path(output_dir)
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"虫害情况报告_{timestamp}.xlsx"
        self.output_path = output_dir / filename
        
        wb.save(self.output_path)
        return True, f"文件已保存:\n{self.output_path}"
    
    def generate_analysis_report(self):
        """生成分析报告"""
        if not self.output_path or not self.output_path.exists():
            return False, "Excel文件不存在"
        
        wb = load_workbook(self.output_path)
        df = pd.read_excel(self.output_path, sheet_name='虫害情况')
        
        if "虫害分析" in wb.sheetnames:
            del wb["虫害分析"]
        ws_report = wb.create_sheet("虫害分析", 0)
        
        total_records = len(df)
        total_pests = df['发现虫害活动'].sum()
        avg_density = total_pests / total_records if total_records > 0 else 0
        max_single = df['发现虫害活动'].max()
        
        pest_type_stats = df.groupby('虫害类型').agg({
            '虫害类型': 'count',
            '发现虫害活动': 'sum'
        }).rename(columns={'虫害类型': '记录数', '发现虫害活动': '总数量'})
        pest_type_stats['占比'] = (pest_type_stats['总数量'] / total_pests * 100).round(1)
        pest_type_stats = pest_type_stats.sort_values('总数量', ascending=False)
        
        building_stats = df.groupby('建筑物').agg({
            '建筑物': 'count',
            '发现虫害活动': 'sum'
        }).rename(columns={'建筑物': '记录数', '发现虫害活动': '总数量'})
        building_stats['占比'] = (building_stats['总数量'] / total_pests * 100).round(1)
        building_stats = building_stats.sort_values('总数量', ascending=False)
        
        top10 = df.nlargest(10, '发现虫害活动')
        
        self._draw_overview_section(ws_report, total_records, total_pests, avg_density, max_single)
        self._draw_pest_type_stats(ws_report, pest_type_stats, 10)
        self._draw_building_stats(ws_report, building_stats, 18 + len(pest_type_stats))
        self._draw_top10_section(ws_report, top10, 26 + len(pest_type_stats) + len(building_stats))
        
        wb.save(self.output_path)
        return True, "分析报告已生成"
    
    def _draw_overview_section(self, ws, total_records, total_pests, avg_density, max_single):
        """绘制数据概览部分"""
        ws.merge_cells('A1:G2')
        title_cell = ws['A1']
        title_cell.value = "虫害情况数据概览"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        overview_data = [
            ('总记录数', f'{total_records} 条'),
            ('虫害总数', f'{total_pests} 只'),
            ('平均密度', f'{avg_density:.1f} 只/处'),
            ('⚠️ 最大单点', f'{max_single} 只')
        ]
        
        row = 4
        for i, (label, value) in enumerate(overview_data):
            col = i * 2 + 1
            
            label_cell = ws.cell(row=row, column=col)
            label_cell.value = label
            label_cell.font = Font(size=11, bold=True)
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            value_cell = ws.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(size=13, bold=True)
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if '⚠️' in label:
                value_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _draw_pest_type_stats(self, ws, pest_type_stats, start_row):
        """绘制虫害类型统计表"""
        row = start_row
        ws.merge_cells(f'A{row}:D{row}')
        subtitle_cell = ws[f'A{row}']
        subtitle_cell.value = "虫害类型统计"
        subtitle_cell.font = Font(size=13, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
        subtitle_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
        headers = ['虫害类型', '记录数', '总数量（只）', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        for pest_type, data in pest_type_stats.iterrows():
            row += 1
            values = [pest_type, int(data['记录数']), int(data['总数量']), f"{data['占比']}%"]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def _draw_building_stats(self, ws, building_stats, start_row):
        """绘制建筑物统计表"""
        row = start_row
        ws.merge_cells(f'A{row}:D{row}')
        subtitle_cell = ws[f'A{row}']
        subtitle_cell.value = "建筑物虫害统计"
        subtitle_cell.font = Font(size=13, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
        subtitle_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
        headers = ['建筑物', '记录数', '总数量（只）', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        for building, data in building_stats.iterrows():
            row += 1
            values = [building, int(data['记录数']), int(data['总数量']), f"{data['占比']}%"]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def _draw_top10_section(self, ws, top10_df, start_row):
        """绘制高危区域TOP10表"""
        row = start_row
        ws.merge_cells(f'A{row}:G{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "高危区域分析 - TOP 10"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        title_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
        headers = ['排名', '建筑物', '楼层', '部门', '监测点位', '虫害类型', '数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        for idx, record in top10_df.iterrows():
            row += 1
            rank = row - start_row - 1
            values = [
                rank,
                record['建筑物'],
                record['楼层'],
                record['部门'],
                record['检查/发现监测点位'],
                record['虫害类型'],
                int(record['发现虫害活动'])
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                if rank <= 3:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['G'].width = 10


class PestReportApp(App):
    """Kivy应用主类"""
    
    def build(self):
        """构建应用界面"""
        Window.clearcolor = (0.95, 0.95, 0.95, 1)
        
        self.extractor = PestReportExtractor()
        self.selected_pdf = None
        
        # 延迟请求权限(避免启动时闪退)
        if ANDROID:
            Clock.schedule_once(self.request_android_permissions, 0.5)
        
        # 主布局
        main_layout = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        
        # 标题
        title = Label(
            text='🐛 虫害报告提取工具',
            size_hint=(1, 0.1),
            font_size=dp(24),
            bold=True,
            color=(0.2, 0.2, 0.2, 1)
        )
        main_layout.add_widget(title)
        
        # 版本信息
        version_label = Label(
            text='v2.0 Android版',
            size_hint=(1, 0.05),
            font_size=dp(14),
            color=(0.5, 0.5, 0.5, 1)
        )
        main_layout.add_widget(version_label)
        
        # 选择文件按钮
        self.select_btn = Button(
            text='📁 选择 PDF 文件',
            size_hint=(1, 0.12),
            font_size=dp(18),
            background_color=(0.3, 0.6, 1, 1),
            background_normal=''
        )
        self.select_btn.bind(on_press=self.select_pdf)
        main_layout.add_widget(self.select_btn)
        
        # 文件名显示
        self.file_label = Label(
            text='未选择文件',
            size_hint=(1, 0.08),
            font_size=dp(14),
            color=(0.4, 0.4, 0.4, 1)
        )
        main_layout.add_widget(self.file_label)
        
        # 处理按钮
        self.process_btn = Button(
            text='🚀 开始处理',
            size_hint=(1, 0.12),
            font_size=dp(18),
            background_color=(0.2, 0.7, 0.3, 1),
            background_normal='',
            disabled=True
        )
        self.process_btn.bind(on_press=self.process_pdf)
        main_layout.add_widget(self.process_btn)
        
        # 进度条
        self.progress = ProgressBar(max=100, size_hint=(1, 0.05))
        main_layout.add_widget(self.progress)
        
        # 状态显示区域
        scroll = ScrollView(size_hint=(1, 0.45))
        self.status_label = Label(
            text='等待操作...\n\n使用说明：\n1. 点击"选择PDF文件"\n2. 选择虫害报告PDF\n3. 点击"开始处理"\n4. 等待生成完成',
            size_hint_y=None,
            font_size=dp(14),
            color=(0.3, 0.3, 0.3, 1),
            halign='left',
            valign='top'
        )
        self.status_label.bind(texture_size=self.status_label.setter('size'))
        scroll.add_widget(self.status_label)
        main_layout.add_widget(scroll)
        
        return main_layout
    
    def request_android_permissions(self, dt):
        """请求Android权限（延迟执行）"""
        try:
            from android import api_version
            
            permissions = [
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE
            ]
            
            # Android 11+ (API 30+) 特殊处理
            if api_version >= 30:
                self.update_status('📱 Android 11+ 检测到\n\n文件将保存到应用专属目录：\n/Android/data/.../files/Documents/虫害报告\n\n无需额外权限！')
                
                # 尝试请求 MANAGE_EXTERNAL_STORAGE（可选）
                try:
                    from jnius import autoclass
                    Intent = autoclass('android.content.Intent')
                    Settings = autoclass('android.provider.Settings')
                    Uri = autoclass('android.net.Uri')
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    
                    # 检查是否有所有文件访问权限
                    if api_version >= 30:
                        Environment = autoclass('android.os.Environment')
                        if not Environment.isExternalStorageManager():
                            # 引导用户到设置页面
                            intent = Intent(Settings.ACTION_MANAGE_APP_ALL_FILES_ACCESS_PERMISSION)
                            uri = Uri.parse(f"package:{PythonActivity.mActivity.getPackageName()}")
                            intent.setData(uri)
                            PythonActivity.mActivity.startActivity(intent)
                            self.update_status('📱 请在设置中授予"所有文件访问权限"\n\n（可选，用于访问共享存储）')
                except Exception as e:
                    print(f"无法请求 MANAGE_EXTERNAL_STORAGE: {e}")
            else:
                # Android 10 及以下
                request_permissions(permissions)
                self.update_status('✅ 权限请求已发送\n如果未弹出权限对话框，请手动在设置中授权')
                
        except Exception as e:
            self.update_status(f'📂 使用应用专属存储\n文件将保存到:\n{STORAGE_PATH}')
    
    def select_pdf(self, instance):
        """选择PDF文件"""
        content = BoxLayout(orientation='vertical', spacing=dp(10))
        
        file_chooser = FileChooserListView(
            path=STORAGE_PATH,
            filters=['*.pdf']
        )
        content.add_widget(file_chooser)
        
        btn_layout = BoxLayout(size_hint=(1, 0.15), spacing=dp(10))
        
        select_btn = Button(text='选择', background_color=(0.2, 0.7, 0.3, 1))
        cancel_btn = Button(text='取消', background_color=(0.8, 0.3, 0.3, 1))
        
        popup = Popup(
            title='选择PDF文件',
            content=content,
            size_hint=(0.9, 0.9)
        )
        
        def on_select(instance):
            if file_chooser.selection:
                self.selected_pdf = file_chooser.selection[0]
                self.file_label.text = f'已选择: {Path(self.selected_pdf).name}'
                self.process_btn.disabled = False
                self.update_status(f'✅ 已选择文件:\n{Path(self.selected_pdf).name}')
            popup.dismiss()
        
        def on_cancel(instance):
            popup.dismiss()
        
        select_btn.bind(on_press=on_select)
        cancel_btn.bind(on_press=on_cancel)
        
        btn_layout.add_widget(select_btn)
        btn_layout.add_widget(cancel_btn)
        content.add_widget(btn_layout)
        
        popup.open()
    
    def process_pdf(self, instance):
        """处理PDF文件"""
        if not self.selected_pdf:
            self.show_message('错误', '请先选择PDF文件')
            return
        
        self.process_btn.disabled = True
        self.select_btn.disabled = True
        self.progress.value = 0
        self.update_status('⏳ 开始处理...')
        
        # 在后台线程执行
        thread = threading.Thread(target=self.extract_and_generate)
        thread.start()
    
    def extract_and_generate(self):
        """提取数据并生成Excel（后台线程）"""
        # 步骤1：提取数据
        Clock.schedule_once(lambda dt: self.update_status('📄 正在读取PDF...'), 0)
        Clock.schedule_once(lambda dt: setattr(self.progress, 'value', 20), 0)
        
        success, message = self.extractor.extract_pest_data_from_pdf(self.selected_pdf)
        
        if not success:
            Clock.schedule_once(lambda dt: self.update_status(f'❌ {message}'), 0)
            Clock.schedule_once(lambda dt: self.enable_buttons(), 0)
            return
        
        Clock.schedule_once(lambda dt: self.update_status(f'✅ {message}'), 0)
        Clock.schedule_once(lambda dt: setattr(self.progress, 'value', 50), 0)
        
        # 步骤2：生成Excel
        Clock.schedule_once(lambda dt: self.update_status('📊 正在生成Excel...'), 0)
        
        success, message = self.extractor.create_excel()
        
        if not success:
            Clock.schedule_once(lambda dt: self.update_status(f'❌ {message}'), 0)
            Clock.schedule_once(lambda dt: self.enable_buttons(), 0)
            return
        
        Clock.schedule_once(lambda dt: self.update_status(f'✅ {message}'), 0)
        Clock.schedule_once(lambda dt: setattr(self.progress, 'value', 75), 0)
        
        # 步骤3：生成分析报告
        Clock.schedule_once(lambda dt: self.update_status('📈 正在生成分析报告...'), 0)
        
        success, message = self.extractor.generate_analysis_report()
        
        if not success:
            Clock.schedule_once(lambda dt: self.update_status(f'⚠️ {message}'), 0)
        else:
            Clock.schedule_once(lambda dt: self.update_status(f'✅ {message}'), 0)
        
        Clock.schedule_once(lambda dt: setattr(self.progress, 'value', 100), 0)
        
        # 完成
        final_msg = f'🎉 处理完成！\n\n提取记录: {len(self.extractor.pest_data)} 条\n保存位置:\n{self.extractor.output_path}'
        Clock.schedule_once(lambda dt: self.update_status(final_msg), 0)
        Clock.schedule_once(lambda dt: self.show_message('完成', final_msg), 0)
        Clock.schedule_once(lambda dt: self.enable_buttons(), 0)
    
    def update_status(self, text):
        """更新状态显示"""
        self.status_label.text = text
    
    def enable_buttons(self):
        """启用按钮"""
        self.process_btn.disabled = False
        self.select_btn.disabled = False
    
    def show_message(self, title, message):
        """显示消息弹窗"""
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(10))
        
        msg_label = Label(
            text=message,
            size_hint=(1, 0.8),
            font_size=dp(14),
            halign='left',
            valign='top'
        )
        msg_label.bind(size=msg_label.setter('text_size'))
        content.add_widget(msg_label)
        
        close_btn = Button(
            text='确定',
            size_hint=(1, 0.2),
            background_color=(0.3, 0.6, 1, 1)
        )
        
        popup = Popup(
            title=title,
            content=content,
            size_hint=(0.8, 0.6)
        )
        
        close_btn.bind(on_press=popup.dismiss)
        content.add_widget(close_btn)
        
        popup.open()


if __name__ == '__main__':
    import sys
    import traceback
    
    try:
        print("=" * 50)
        print("🐛 虫害报告提取工具 v2.0")
        print(f"Android: {ANDROID}")
        print(f"Storage Path: {STORAGE_PATH}")
        print(f"Python Version: {sys.version}")
        print("=" * 50)
        
        app = PestReportApp()
        app.run()
    except Exception as e:
        error_msg = f"\n\n应用启动失败:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        
        # 尝试写入错误日志
        try:
            error_file = Path(STORAGE_PATH) / "pest_error.log"
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(error_msg)
            print(f"\n错误日志已保存: {error_file}")
        except:
            pass
        
        raise
